"""
Script to evaluate the quality of LLM-as-a-judge by manually comparing its judgments with human evaluations.

This script:
1. Loads model generation results from 3 datasets (WikiTQ, WikiSQL, FetaQA)
2. Randomly samples 100 examples
3. Runs LLM-as-a-judge evaluation
4. Creates a spreadsheet with results for manual human evaluation
"""

import argparse
import json
import random
from pathlib import Path
from typing import Any, Dict, List

import dotenv
import openpyxl
from openpyxl.styles import Alignment, Font
from tqdm import tqdm

from generation.generator_vllm_gemma3 import VLLMGeneratorGemma3


dotenv.load_dotenv("../.env")


def load_json_with_prefix(filepath: str, prefix: str) -> Dict[str, Any]:
    """Load a JSON file and add a prefix to all keys."""
    with open(filepath, "r") as f:
        data = json.load(f)

    # Add prefix to keys
    prefixed_data = {f"{prefix}_{key}": value for key, value in data.items()}
    return prefixed_data


def combine_datasets(wikitq_path: str, wikisql_path: str) -> Dict[str, Any]:
    """Load and combine the three datasets with appropriate prefixes."""
    print("Loading datasets...")

    wikitq_data = load_json_with_prefix(wikitq_path, "wikitq")
    print(f"  Loaded {len(wikitq_data)} examples from WikiTQ")

    wikisql_data = load_json_with_prefix(wikisql_path, "wikisql")
    print(f"  Loaded {len(wikisql_data)} examples from WikiSQL")

    # Combine all datasets
    combined_data = {**wikitq_data, **wikisql_data}
    print(f"Total combined examples: {len(combined_data)}")

    return combined_data


def sample_examples(combined_data: Dict[str, Any], n_samples: int = 50, seed: int = 42) -> Dict[str, Any]:
    """Randomly sample n examples from the combined dataset."""
    random.seed(seed)

    all_keys = list(combined_data.keys())
    sampled_keys = random.sample(all_keys, min(n_samples, len(all_keys)))

    sampled_data = {key: combined_data[key] for key in sampled_keys}
    print(f"Sampled {len(sampled_data)} examples")

    return sampled_data


def prepare_llm_judge_requests(sampled_data: Dict[str, Any]) -> tuple[List[Dict[str, Any]], List[str]]:
    """Prepare batch requests for LLM-as-a-judge evaluation."""
    batch_requests = []
    example_ids = []

    for example_id, output in tqdm(sampled_data.items(), desc="Preparing LLM-as-a-judge requests"):
        # Extract prediction from "Step 2:"
        if "Step 2 :" in output["generations"][0]:
            pred_answer = output["generations"][0].split("Step 2 :")[-1].strip()
        else:
            pred_answer = output["generations"][0].split("Step 2:")[-1].strip()

        # Extract gold answer and other metadata
        gold_answer = output["ori_data_item"]["answer_text"]
        question = output["ori_data_item"]["question"]
        page_title = output["ori_data_item"]["table_with_metadata"]["page_title"]
        section_title = output["ori_data_item"]["table_with_metadata"].get("section_title", "")

        batch_requests.append(
            {
                "oracle_answer_list": gold_answer,
                "our_answer": pred_answer,
                "question": question,
                "page_title": page_title,
                "section_title": section_title,
            }
        )
        example_ids.append(example_id)

    return batch_requests, example_ids


def create_evaluation_spreadsheet(
    sampled_data: Dict[str, Any],
    example_ids: List[str],
    llm_judge_scores: List[int],
    output_path: str,
    image_base_path: Path,
):
    """
    Create a spreadsheet with evaluation results for manual human evaluation.

    Columns: ID, Question, Given Answer, Gold Answer, LLM-as-a-judge Score, Human Judgment
    """
    print("\nCreating evaluation spreadsheet...")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LLM-as-a-Judge Evaluation"

    # Set column widths
    ws.column_dimensions["A"].width = 15  # ID
    ws.column_dimensions["B"].width = 40  # Question
    ws.column_dimensions["C"].width = 40  # Given Answer
    ws.column_dimensions["D"].width = 40  # Gold Answer
    ws.column_dimensions["E"].width = 15  # LLM Score
    ws.column_dimensions["F"].width = 30  # Human Judgment

    # Header row
    headers = ["ID", "Question", "Given Answer", "Gold Answer", "LLM Score", "Human Agrees?"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add data rows
    for idx, (example_id, score) in enumerate(
        tqdm(zip(example_ids, llm_judge_scores), desc="Adding rows to spreadsheet")
    ):
        row_num = idx + 2  # +2 because header is row 1 and we start from row 2

        output = sampled_data[example_id]

        # Extract data
        question = output["ori_data_item"]["question"]
        gold_answer = output["ori_data_item"]["answer_text"]
        gold_answer_str = gold_answer[0] if isinstance(gold_answer, list) else str(gold_answer)

        # Extract prediction
        if "Step 2 :" in output["generations"][0]:
            pred_answer = output["generations"][0].split("Step 2 :")[-1].strip()
        else:
            pred_answer = output["generations"][0].split("Step 2:")[-1].strip()

        # Create row data
        ws.cell(row=row_num, column=1, value=example_id)
        ws.cell(row=row_num, column=2, value=question)
        ws.cell(row=row_num, column=3, value=pred_answer)
        ws.cell(row=row_num, column=4, value=gold_answer_str)
        ws.cell(row=row_num, column=5, value="CORRECT" if score == 1 else "INCORRECT")
        ws.cell(row=row_num, column=6, value="")  # Empty for human to fill

        # Wrap text for long answers
        ws.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    # Save workbook
    wb.save(output_path)
    print(f"Spreadsheet saved to: {output_path}")


def get_gemma3_generator(
    model_name: str = "google/gemma-3-27b-it",
    prompt_mode: str = "text-image",
    number_of_gpus: int = 2,
    limit_mm_per_prompt: dict = None,
    system_prompt: str = None,
    context_window_size: int = 10000,
):
    args = argparse.Namespace()
    args.engine = "vllm"
    args.vllm_model_name = model_name
    args.number_of_gpus = number_of_gpus
    args.max_api_total_tokens = context_window_size
    args.prompt_mode = prompt_mode
    args.prompt_style = ""  # not needed for image captioning (but if not specified we get an error)
    args.seed = 42  # also not needed (but if not specified we get an error)

    generator = VLLMGeneratorGemma3(
        args,
        limit_mm_per_prompt=limit_mm_per_prompt,
        system_prompt=system_prompt,
        gpu_memory_utilization=0.9,
    )
    return generator


def main():
    parser = argparse.ArgumentParser(description="Evaluate LLM-as-a-judge quality")
    parser.add_argument(
        "--wikitq_path",
        type=str,
        default=".CAPTR/results/gemma3-mmtabqa-captioned-short-interleaved-reasoner/wikitq_AQ/interleaved_reasoning_output.json",
        help="Path to WikiTQ AQ results",
    )
    parser.add_argument(
        "--wikisql_path",
        type=str,
        default="CAPTR/results/gemma3-mmtabqa-captioned-short-interleaved-reasoner/wikisql_AQ/interleaved_reasoning_output.json",
        help="Path to WikiSQL AQ results",
    )
    parser.add_argument(
        "--output_path",
        type=str,
        default="CAPTR/statistics/llm_judge_quality_evaluation.xlsx",
        help="Output path for the evaluation spreadsheet",
    )
    parser.add_argument(
        "--image_base_path",
        type=str,
        default="CAPTR/datasets/MMTabQA_Dataset/IMAGES",
        help="Base path for images",
    )
    parser.add_argument("--n_samples", type=int, default=100, help="Number of examples to sample (default: 100)")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for sampling")
    parser.add_argument(
        "--validate_only", action="store_true", help="Only validate paths and load data without running LLM-as-a-judge"
    )

    args = parser.parse_args()

    # 1. Load and combine datasets
    combined_data = combine_datasets(args.wikitq_path, args.wikisql_path)

    # 2. Sample examples
    sampled_data = sample_examples(combined_data, n_samples=args.n_samples, seed=args.seed)

    # 3. Prepare LLM-as-a-judge requests
    batch_requests, example_ids = prepare_llm_judge_requests(sampled_data)

    print("\nSample distribution:")
    wikitq_count = sum(1 for eid in example_ids if eid.startswith("wikitq_"))
    wikisql_count = sum(1 for eid in example_ids if eid.startswith("wikisql_"))
    fetaqa_count = sum(1 for eid in example_ids if eid.startswith("fetaqa_"))
    print(f"  WikiTQ: {wikitq_count}")
    print(f"  WikiSQL: {wikisql_count}")
    print(f"  FetaQA: {fetaqa_count}")

    if args.validate_only:
        print("\n✅ Validation complete! Data loaded successfully.")
        print("Remove --validate_only flag to run full evaluation with LLM-as-a-judge.")
        return

    # 4. Load generator and run LLM-as-a-judge
    print("\nInitializing LLM-as-a-judge model...")

    generator = get_gemma3_generator()

    print(f"\nEvaluating {len(batch_requests)} examples using LLM-as-a-judge...")
    llm_judge_scores = generator.evaluate_llm_as_a_judge(batch_requests)

    # store llm_judge_scores in a json file
    with open("llm_judge_scores.json", "w") as f:
        json.dump(llm_judge_scores, f)

    # 5. Create evaluation spreadsheet
    image_base_path = Path(args.image_base_path)
    create_evaluation_spreadsheet(sampled_data, example_ids, llm_judge_scores, args.output_path, image_base_path)

    print(
        "\n✅ Done! Please open the spreadsheet and manually evaluate whether you agree with the LLM-as-a-judge judgments."
    )


if __name__ == "__main__":
    main()
